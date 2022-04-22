import requests
import pprint
import json
from openpyxl import Workbook
from openpyxl.styles import Font

gdr_endpoint = "https://apigw.intra-dev01.bdf-dev01.local:443/gateway/SDL_GestionDesRemettants/1.0.0"

headers = {
    "accept": "application/json",
    "x-Gateway-APIKey": "0ab73db6-81c2-44ac-9a83-9ec979a2b97d"
}

def idGdrFromCib(date, cib):
    params = {
        "date": date,
        "cib": cib
    }
    s = requests.Session()
    r = s.request("GET", f"{gdr_endpoint}/etablissements", headers=headers, params=params, verify=False)
    s.close()
    resp = r.json()
    return resp["items"][0]["idGDR"]

def idGdrFromSiren(date, siren):
    params = {
        "date": date,
        "siren": siren
    }
    s = requests.Session()
    r = s.request("GET", f"{gdr_endpoint}/etablissements", headers=headers, params=params, verify=False)
    s.close()
    resp = r.json()
    if resp["items"]:
        return resp["items"][0]["idGDR"]
    return None

def donneesEtablissementFromIdGdr(date, idGdr):
    params = {
        "date": date,
        "idGdr": idGdr
    }
    s = requests.Session()
    r = s.request("GET", f"{gdr_endpoint}/etablissements/{idGdr}/donnees", headers=headers, params=params, verify=False)
    s.close()
    resp = r.json()
    return resp

def historiqueFromCodeDonnee(date, idGdr, codeDonnee):
    params = {
        "date": date,
        "idGdr": idGdr
    }
    s = requests.Session()
    r = s.request("GET", f"{gdr_endpoint}/etablissements/{idGdr}/historique/{codeDonnee}", headers=headers, params=params, verify=False)
    s.close()
    resp = r.json()
    return resp

def attendusFromIdGdr(startDate, endDate, idGdr):
    params = {
        "startDate": startDate,
        "endDate": endDate,
        "idGdr": idGdr
    }
    s = requests.Session()
    r = s.request("GET", f"{gdr_endpoint}/etablissements/{idGdr}/distinctAttendus", headers=headers, params=params, verify=False)
    s.close()
    resp = r.json()
    return resp

# TODO

def donneesEtablissementFromCib(date, cib):
    idGdr = idGdrFromCib(date, cib)
    return donneesEtablissementFromIdGdr(date, idGdr)

def donneesEtablissementFromSiren(date, siren):
    idGdr = idGdrFromSiren(date, siren)
    return donneesEtablissementFromIdGdr(date, idGdr)

def attendusFromCib(startDate, endDate, cib):
    idGdr = idGdrFromCib(startDate, cib)
    return attendusFromIdGdr(startDate, endDate, idGdr)

def attendusFromSiren(startDate, endDate, siren):
    idGdr = idGdrFromSiren(startDate, siren)
    return attendusFromIdGdr(startDate, endDate, idGdr)

def autosizeCells(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

def gdrToExcel():
    # import config
    with open('config_test.json', 'r', encoding='utf-8') as f:
        conf = json.load(f)
        queries = conf["gdr"]

    for query in queries:
        print(list(query.keys()))
        if "cib" in list(query.keys()):
            ID_LABEL = "CIB"
            ID_VALUE = query["cib"]
            IDGDR = idGdrFromCib(query["date"], ID_VALUE)
            donneesCV = donneesEtablissementFromCib(query["date"], query["cib"])
            attendus = attendusFromCib(query["startDate"], query["endDate"], query["cib"])
        elif "siren" in list(query.keys()):
            ID_LABEL = "SIREN"
            ID_VALUE = query["siren"]
            IDGDR = idGdrFromSiren(query["date"], ID_VALUE)
            donneesCV = donneesEtablissementFromSiren(query["date"], query["siren"])
            attendus = attendusFromSiren(query["startDate"], query["endDate"], query["siren"])
        else:
            print("[!] Entrer un identifiant")

        print(f"[!] Current {ID_LABEL} : {ID_VALUE}")

        columnLabels = list(donneesCV[0].keys())

        # creation des onglets
        wb = Workbook()
        del wb['Sheet']
        ws0 = wb.create_sheet("Paramètres")
        ws0.title = "Paramètres"
        ws1 = wb.active
        ws1 = wb.create_sheet("Donnees_CV")
        ws1.title = "Donnees_CV"
        ws2 = wb.create_sheet("Attendus")
        ws2.title = "Attendus"
        ws3 = wb.create_sheet("Historiques")
        ws3.title = "Historiques"

        # filling table
        parameterLabels = [ID_LABEL, "date", "startDate", "endDate"]
        ws0.append(parameterLabels)
        ws0.append([ID_VALUE, query["date"], query["startDate"], query["endDate"]])

        ws1.append(columnLabels)
        for donnee in donneesCV:
            ws1.append(list(donnee.values()))

        columnLabels = list(attendus[0].keys())[1:]
        ws2.append(columnLabels)
        for attendu in attendus:
            ws2.append(list(attendu.values())[1:])

        h = historiqueFromCodeDonnee(query["date"], IDGDR, donneesCV[0]["codeDonneeDictionnaire"])["donnees"]
        historiqueLabel = list(h[0].keys())
        ws3.append(historiqueLabel)
        for donnee in donneesCV:
            historiques = historiqueFromCodeDonnee(query["date"], IDGDR, donnee["codeDonneeDictionnaire"])["donnees"]
            for histo in historiques:
                ws3.append(list(histo.values()))

        # applying styles
        for cell in list(ws0.rows)[0]:
            cell.style = 'Headline 3'
            cell.font = Font(b=True, color="000000")

        for cell in list(ws1.rows)[0]:
            cell.style = 'Headline 3'
            cell.font = Font(b=True, color="000000")

        for cell in list(ws2.rows)[0]:
            cell.style = 'Headline 3'
            cell.font = Font(b=True, color="000000")

        for cell in list(ws3.rows)[0]:
            cell.style = 'Headline 3'
            cell.font = Font(b=True, color="000000")

        autosizeCells(ws0)
        autosizeCells(ws1)
        autosizeCells(ws2)
        autosizeCells(ws3)

        ws1.auto_filter.ref = ws1.dimensions
        ws2.auto_filter.ref = ws2.dimensions
        ws3.auto_filter.ref = ws2.dimensions

        wb.save(f"{ID_LABEL}_{ID_VALUE}.xlsx")
        print(f"[*] Document {ID_LABEL}_{ID_VALUE}.xlsx saved to disk\n")